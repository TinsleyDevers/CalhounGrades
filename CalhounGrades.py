from selenium import webdriver
from selenium.webdriver.common.by import By
from bs4 import BeautifulSoup
from openpyxl.styles import PatternFill, Font, Border, Side, Alignment
import openpyxl
import pandas as pd
import time
import re

# set your username and password here!!!
username = "Your Calhoun Email Here (e.g. yourname@calhoun.edu)"
password = "Your Password Here!"
    
# if you get any errors, try changing time.sleep values or changing the exclude list

# exits if you didn't set your username and password above.
if username == "Your Calhoun Email Here (e.g. yourname@calhoun.edu)" or password == "Your Password Here!":
    print("Please set your username and password in the script.")
    exit()

# calculation logic
def calculate_percentage(points):
    match = re.match(r"([\d,]+\.?\d*) / ([\d,]+\.?\d*)", points.replace(',', ''))
    if match:
        numerator, denominator = map(float, match.groups())
        return (numerator / denominator) * 100
    return None

def is_percentage(grade):
    return bool(re.match(r"^\d+(\.\d+)?%$", grade.strip()))

def determine_letter_grade(percent):
    if percent >= 90:
        return 'A'
    elif percent >= 80:
        return 'B'
    elif percent >= 70:
        return 'C'
    elif percent >= 60:
        return 'D'
    else:
        return 'F'

# remove dupes logic
def remove_duplicates(lst):
    seen = set()
    seen_add = seen.add
    return [x for x in lst if not (x in seen or seen_add(x))]

driver = webdriver.Chrome()

# calhoun website login
driver.get("https://id.quicklaunch.io/calhoun")
time.sleep(2)

# inputs username in username field
username_field = driver.find_element(By.ID, "usernameUserInput")
username_field.send_keys(username)

# inputs password in password field
password_field = driver.find_element(By.ID, "password")
password_field.send_keys(password)

# clicks login button
login_button = driver.find_element(By.CLASS_NAME, "form-actions")
login_button.click()
time.sleep(5)

main_window = driver.current_window_handle

# launch blackboard to save login data to later open the grade page
blackboardlink = driver.find_element(By.XPATH, "//a[@href='https://blackboard.calhoun.edu/auth-saml/saml/login?apId=_192_1&redirectUrl=https%3A%2F%2Fblackboard.calhoun.edu%2Fultra%2Finstitution-page']")
blackboardlink.click()
time.sleep(4)

# close new window
driver.switch_to.window(driver.window_handles[1])
driver.close()
driver.switch_to.window(main_window)

# open grades page blackboard
driver.get("https://blackboard.calhoun.edu/ultra/grades")
time.sleep(3)

# scroll to load all grades
div = driver.find_element(By.ID, 'main-content-inner')
for _ in range(5):
    driver.execute_script("arguments[0].scrollBy(0, window.innerHeight);", div)
time.sleep(3)

soup = BeautifulSoup(driver.page_source, 'html.parser')

# if getting different numbers error exclude courses that don't have a grade.
exclude = ['Blackboard Student Orientation (BSO)', 'Student Math Resources', 'Student Resources', 'Work-Based Learning Modules']

# courses and grades extracted from their elements
courses = [element.text for element in soup.select('.subheader .bb-click-target') if element.text not in exclude]
grades = [a.find('span', class_='grade-input-display grade-ellipsis').find('bdi').text for a in soup.find_all('a') if a.get('bb-peek-sref') == '::baseGradesStudent.navigateToGradePanel()']

courses = remove_duplicates(courses)
grades = remove_duplicates(grades)

# courses with grades and filter based on list
courses_with_grades = list(zip(courses, grades))
filtered_courses_with_grades = [(course, grade) for course, grade in courses_with_grades if course not in exclude]

# turn back into individual lists
filtered_courses = [item[0] for item in filtered_courses_with_grades]
filtered_grades = [item[1] for item in filtered_courses_with_grades]

# check if already has a % grade if so just use that % for percent and letter and N/A for points
processed_data = []
for grade in filtered_grades:
    if is_percentage(grade):
        percent = float(grade.strip('%'))
        points = "N/A"
        letter = determine_letter_grade(percent)
    else:
        points = grade if re.match(r"[\d,]+\.?\d* / [\d,]+\.?\d*", grade) else "N/A"
        percent = calculate_percentage(grade) if points != "N/A" else None
        letter = determine_letter_grade(percent) if percent is not None else None
    
    processed_data.append((points, percent if percent is not None else grade, letter))

# use the processed data in df
df = pd.DataFrame({'Course': filtered_courses, 'Points': [data[0] for data in processed_data], 'Percent': [data[1] for data in processed_data], 'Letter': [data[2] for data in processed_data]})

# widths and file save name
column_widths = {'Course': 530, 'Points': 160, 'Percent': 100, 'Letter': 100}
filename = 'grades.xlsx'

with pd.ExcelWriter(filename, engine='xlsxwriter') as writer:
    df.to_excel(writer, index=False)

    workbook = writer.book
    worksheet = writer.sheets['Sheet1']
print('Preparing for formatting...')
print('Please wait...')
print('-----------------')

# open using openpyxl to format
book = openpyxl.load_workbook(filename)
sheet = book.active

# styles
darkbg = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')
lightbg = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')
borderstyle = Side(border_style="thin")
border = Border(top=borderstyle, bottom=borderstyle)

# bold headers
for cell in sheet[1]:
    cell.font = Font(bold=True)

# background colors
for row in sheet.iter_rows(min_row=2):
    for cell in row:
        if cell.row % 2 == 0:
            cell.fill = darkbg
        else:
            cell.fill = lightbg

# borders & center alignment
for row in sheet.iter_rows():
    for cell in row:
        cell.border = border
        cell.alignment = Alignment(horizontal='center')

# column widths using column_widths from before
for column, width in column_widths.items():
    col_idx = df.columns.get_loc(column) + 1
    sheet.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = width / 8

# save formatting
book.save(filename)

# DONE!
print('Successfully formatted!')
print(f'Saving as {filename}')
time.sleep(1)
driver.quit()
print('Successfully saved!')